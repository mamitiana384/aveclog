import streamlit as st
import streamlit as st
import pandas as pd
import io
import re
import time  # Pour simuler des chargements
from io import BytesIO
from reportlab.lib.pagesizes import letter # type: ignore
from reportlab.pdfgen import canvas # type: ignore
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import streamlit as st
from io import BytesIO
import plotly.express as px
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode # type: ignore
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
# Function to standardize the values for consistent searching
# Fonction de normalisation pour les numéros de téléphone
# Vérification de la connexion
if "authenticated" not in st.session_state or not st.session_state["authenticated"]:
    st.warning("🔴 Vous devez vous connecter.")
    st.switch_page("pages/login.py")
    st.stop()

# Interface principale
st.title(f"Bienvenue {st.session_state['username']} 👋")
st.write("Vous êtes connecté avec succès.")
# Stylisation du bouton de déconnexion avec le CSS que tu as fourni
st.markdown("""
    <style>
        .btn-grad {
            background-image: linear-gradient(to right, #FC354C 0%, #0ABFBC 51%, #FC354C 100%);
            margin: 10px;
            padding: 15px 45px;
            text-align: center;
            text-transform: uppercase;
            transition: 0.5s;
            background-size: 200% auto;
            color: white;            
            box-shadow: 0 0 20px #eee;
            border-radius: 10px;
            display: block;
        }

        .btn-grad:hover {
            background-position: right center; /* change the direction of the change here */
            color: #fff;
            text-decoration: none;
        }
    </style>
""", unsafe_allow_html=True)

# Bouton avec la classe personnalisée
if st.button("🔓", key="logout", help="Se déconnecter", use_container_width=True):
    # Appliquer le style personnalisé
    st.markdown('<a href="#" class="btn-grad">Déconnexion</a>', unsafe_allow_html=True)
    st.session_state["authenticated"] = False
    st.switch_page("pages/login.py")  # Retour à la page de connexion

def normalize_value(value):
    # Supprimer tous les caractères non numériques pour les numéros
    return re.sub(r'\D', '', value).lstrip('0')
# Liste des Fokotany fournie par l'utilisateur
fokotany_list = [
    "AMBATOMENA AMBOASARIKELY", "AMPANDRANA ANKADIVATO", "AMPARIBE AMBOHIDAHY MAHAMASINA", "ANATIHAZO ISOTRY", 
    "ANDAVAMAMBA ANATIHAZO I", "ANDAVAMAMBA ANATIHAZO II", "ANDAVAMAMBA ANJEZIKA II", "ANDRANOMANALINA AFOVOANY",
    "ANDRANOMANALINA 1", "ANDRANOMANALINA ISOTRY", "ANTETEZANAFOVOANY 1", "ANTETEZANAFOVOANY 2", "ANTOHOMADINIKA AFOVOANY",
    "ANTOHOMADINIKA AVARATRA ANTANI", "67HA AFOVOANY ANDREFANA", "CITE 67 HA ATSIMO", "67 HA AVARATRA ANDREFANA", 
    "CITE AMBODIN_ISOTRY", "CITE AMPEFILOHA", "FARAVOHITRA MANDROSOA", "ISORAKA AMPATSAKANA", "ISOTRY ATSINANANA TSARALALANA (FIATA)",
    "MANARINTSOA AFOVOANY", "MANARINTSOA ATSINANANA", "SOARANO AMBONDRONA AMBODIFILAO", "SOARANO AMBONDRONA TSIAZOTAFO", 
    "AMBATOMITSANGANA", "AMBOHITRAKELY", "AMPANDRANA ATSINANANA", "AMPANDRANA BESARETY", "ANKADIFOTSY ANTANIFOTSY", 
    "ANKADIFOTSY BEFELATANANA", "ANKAZOMANGA ANDRAHARO", "BEHORIRIKA AMBATOMITSANGANA", "BESARETY", "BETONGOLO", "MANDIALAZA ANKADIFOTSY", 
    "TSARAMASAY", "ALAROBIA AMBONILOHA", "AMBATOBE", "AMBATOKARANANA", "AMBATOMAINTY", "AMBODIVOANJO", "AMBOHIDAHY",
    "AMPANOTOKANA", "ANALAMAHITSY CITE", "ANDROHIBE", "ANJANAHARY IIA", "ANJANAHARY IIN", "ANJANAHARY IIO", "ANJANAHARY IIS",
    "MANJAKARAY IIB", "MANJAKARAY IID", "MORARANO", "AMBATOROKA", "AMBOHIMIANDRA", "AMBOHIPOTSY", "AMBOHITSIROA VN", 
    "AMBOHITSOA", "AMPAMANTANANA", "ANDOHAMANDRY", "ANKAZOTOKANA AMBONY", "ANTANIMORA AMPASANIMALO", "ANTSAHABE", 
    "FALIARIVO AMBANIDIA", "MAHAZOARIVO", "MANAKAMBAHINY", "MANDROSEZA", "MANJAKAMIADANA", "MIANDRARIVO", "MORARANO", 
    "VOLOSARIKA", "AMBANIN-AMPAMARINANA", "AMBOHIBARIKELY", "AMPANGABE ANJANKINIFOLO", "ANDAVAMAMBA AMBILANIBE", "ANKADILALANA",
    "ANOSIBE OUEST I", "ANOSIPATRANA EST", "ANOSIPATRANA OUEST", "FIADANANA IIIL", "FIADANANA IIIN", "ILANIVATO AMPASIKA", 
    "IVOLANIRAY", "MAHAMASINA SUD", "MANANJARA", "OUEST AMBOHIJANAHARY IIIG/IIIM", "SOANIERANA III I", "SOANIERANA III J", 
    "TSARAFARITRA", "TSIMIALONJAFY", "AMBATOLAMPY", "AMBOAVAHY", "AMBODIVONA", "AMBODIVONKELY", "AMBOHIDROA", 
    "AMBOHIMIADANA ATSIMO", "AMBOHIMIADANA AVARATRA", "AMBOHIMITSINJO", "AMPANDRIAMBEHIVAVY", "ANJANAKIMBORO", "ANOSISOA", 
    "ANOSIVAVAKA", "ANTANETY ATSIMO", "ANTANJOMBE AMBONY", "ANTANJOMBE AVARATRA", "ANTSARARAY", "AVARATANANA", 
    "AVARATETEZANA", "BETAFO", "Ankasina", "Antohomadinika Atsimo", "Andohatapenaka II", "Andavamamba Anjezika I", 
    "Andohatapenaka I", "Antohomadinika III G Hangar", "Antohomadinika IFAMII", "Andohatapenaka III", "Ambohipo", 
    "Andohanimandroseza", "Ambolokandrina", "Androndrakely", "Tsiadana", "Ankorondrano Andrefana", "Ankorondrano Andranomahery", 
    "Ampahibe", "Andrahavoahangy Antsinanana", "Avaradoha", "Ouest Mananjara", "Ankaditoho Maroroho", 
    "Angarangarana", "Ankazotoho Anosimasimahavelona", "Ampefiloha Ambodirano", "Anosizato Est I", "Mandrangobato II", 
    "Anosizato Est II", "Madera Namontana", "Ouest Ankadimbahoaka", "Anosibe Ouest II", "Mandrangombato I", "Amboditsiry", 
    "Soavimasoandro", "Analamahitsy Tanana", "Ambatomaro", "Tsarahonenana", "Andraisoro", "Ankerana Ankadindramamy", 
    "Ivandry", "Ambohimirary", "Nanisana Iadiambola", "Manjakaray IIC", "Ambodimita", "Ambohimandroso", "Ankazomanga Atsimo", 
    "Andranomena", "Ambodihady", "Antanety Avaratra"
]

def organize_data(df):
    organized_columns = {
        "RSU": [],
        "Arrondissement": [],
        "Uri": [],
        "Code Menage": [],
        "Nom et Prenom1": [],
        "CIN1": [],
        "Nom et Prenom2": [],
        "CIN2": [],
        "Mobile Recepteur": [],
        "Fokotany": []
    }

    # Fonction pour normaliser les numéros mobiles
    def normalize_mobile_number(value):
        value = re.sub(r"[^\d]", "", value)  # Supprime tout sauf les chiffres
        if re.match(r"^3[23-4]\d{7}$", value):  # Préfixes valides sans le zéro
            return "0" + value
        elif re.match(r"^0[23-4]\d{8}$", value):  # Numéro déjà formaté correctement
            return value
        else:
            return None

    # Fonction pour vérifier si une valeur est un nom valide (pas de chiffres)
    def is_valid_name(value):
        return not bool(re.search(r"\d", value))  # Vérifie l'absence de chiffres

    # Parcourir chaque ligne du DataFrame
    for _, row in df.iterrows():
        rsu = None
        arrondissement = None
        uri = None
        code_menage = None
        nom_prenom1 = None
        cin1 = None
        nom_prenom2 = None
        cin2 = None
        mobile = None
        fokotany = None

        for value in row:
            if pd.isna(value):  # Ignorer les valeurs vides
                continue

            value = str(value).strip()  # Nettoyer la valeur
            value_upper = value.upper()

            # Détecter les différents types de données
            if re.match(r"C3T\d+", value):  # RSU
                rsu = value
            elif re.match(r"\d{1,2}(E|ÈME) ARRONDISSEMENT", value_upper):  # Arrondissement
                arrondissement = value_upper
            elif re.match(r"uuid:[a-f0-9\-]+", value):  # URI
                uri = value
            elif re.match(r"\d{16}", value):  # Code Ménage
                code_menage = value
            elif re.match(r"^\d{12}$", value):  # CIN
                if cin1 is None:
                    cin1 = value
                else:
                    cin2 = value
            elif mobile is None and normalize_mobile_number(value):  # Mobile
                mobile = normalize_mobile_number(value)
            elif value_upper in [f.upper() for f in fokotany_list]:  # Fokotany
                fokotany = value
            elif is_valid_name(value):  # Nom et Prénom
                if nom_prenom1 is None:
                    nom_prenom1 = value
                else:
                    nom_prenom2 = value

        # Ajouter les valeurs dans les colonnes organisées
        organized_columns["RSU"].append(rsu)
        organized_columns["Arrondissement"].append(arrondissement)
        organized_columns["Uri"].append(uri)
        organized_columns["Code Menage"].append(code_menage)
        organized_columns["Nom et Prenom1"].append(nom_prenom1)
        organized_columns["CIN1"].append(cin1)
        organized_columns["Nom et Prenom2"].append(nom_prenom2)
        organized_columns["CIN2"].append(cin2)
        organized_columns["Mobile Recepteur"].append(mobile)
        organized_columns["Fokotany"].append(fokotany)

    # Convertir les données organisées en DataFrame
    organized_df = pd.DataFrame(organized_columns)

    return organized_df

# Fonction pour extraire et normaliser toutes les valeurs saisies
def extract_values(search_values):
    # Séparer les valeurs par espaces, retours à la ligne, virgules, etc.
    raw_values = re.split(r'[,\s\n]+', search_values)
    return [normalize_value(value) for value in raw_values if value]


# Fonction pour le chatbot - Réponses basées sur les actions de l'utilisateur
def chatbot_response(user_input, df, df_transactions):
    response = ""

    if "doublons" in user_input.lower():
        if df.duplicated().sum() > 0:
            response = f"Le fichier contient {df.duplicated().sum()} doublon(s)."
        else:
            response = "Aucun doublon détecté dans le fichier."
    
    elif "transactions multiples" in user_input.lower():
        col_name = st.selectbox("Sélectionnez la colonne dans les données", df.columns)
        col_transactions = st.selectbox("Sélectionnez la colonne correspondante dans les transactions", df_transactions.columns)

        transaction_column = st.selectbox("Sélectionnez la colonne des montants de transaction", df_transactions.columns)
        transactions_multiples = df_transactions.groupby(col_transactions)[transaction_column].count()
        transactions_multiples = transactions_multiples[transactions_multiples > 1]

        if not transactions_multiples.empty:
            response = f"Voici les valeurs avec plusieurs transactions : {transactions_multiples.to_dict()}"
        else:
            response = "Aucune valeur n'a effectué plusieurs transactions."

    elif "bonjour" in user_input.lower():
        response = "Bonjour ! Comment puis-je vous aider avec vos fichiers Excel ? Vous pouvez me demander de détecter des doublons ou des transactions multiples."

    else:
        response = "Désolé, je ne comprends pas cette question. Vous pouvez me demander de vérifier des doublons ou des transactions multiples."

    return response
# Fonction pour uniformiser les textes en supprimant les espaces en trop
def uniformize_text(df, column):
    df[column] = df[column].str.strip()  # Supprimer les espaces au début et à la fin
    df[column] = df[column].str.replace(r'\s+', ' ', regex=True)  # Remplacer les espaces multiples par un seul espace
    return df
# Fonction pour uniformiser les en-têtes
def uniformize_headers(df, target_headers):
    df.columns = [target_headers.get(col, col) for col in df.columns]
    return df
def normalize_column_names(df):
    df.columns = df.columns.str.strip().str.lower()  # Supprimer les espaces et mettre en minuscules
    return df
def nettoyer_donnees(df):
    # Nettoyage des numéros de téléphone
    if 'phone' in df.columns:
        df['phone'] = df['phone'].apply(lambda x: re.sub(r'\s+', '', str(x)))  # Retirer les espaces
        df['phone'] = df['phone'].apply(lambda x: '0' + x if not x.startswith('0') else x)  # Ajouter '0' si nécessaire

    # Mise en forme des montants
    if 'amount' in df.columns:
        df['amount'] = df['amount'].apply(lambda x: re.sub(r'[^\d.]', '', str(x)))  # Garder uniquement les chiffres et le point décimal
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')  # Convertir en numérique

    # Mise en forme des dates
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce', dayfirst=True)  # Conversion en format date
    
    # Suppression des doublons
    df = df.drop_duplicates()

    return df

# Fonction pour détecter les doublons
def apply_header_format(worksheet, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in worksheet[1]:  # Appliquer la couleur aux en-têtes (ligne 1)
        cell.fill = fill
def add_pivot_table(workbook, data_sheet_name, pivot_sheet_name, df):
    # Ajout d'une nouvelle feuille pour le TCD
    pivot_sheet = workbook.create_sheet(title=pivot_sheet_name)
    
    # Ajout du Tableau Croisé Dynamique
    data_range = f"{data_sheet_name}!A1:{openpyxl.utils.get_column_letter(df.shape[1])}{df.shape[0]+1}"
    pivot_table = Table(displayName="PivotTable", ref=data_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    pivot_table.tableStyleInfo = style
    pivot_sheet.add_table(pivot_table)

def detect_duplicates(df, column_name):
    df_filtered = df[df[column_name].notna() & (df[column_name] != '0') & (df[column_name] != '')]
    duplicates = df_filtered[df_filtered.duplicated(subset=[column_name], keep=False)]
    return duplicates
def generate_report(recap_df, merged_df, file1_name, file2_name):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Titre du rapport
    c.setFont("Helvetica-Bold", 16)
    c.drawString(100, height - 50, "Rapport de Croisement de Fichiers")

    # Détails des fichiers
    c.setFont("Helvetica", 12)
    c.drawString(100, height - 100, f"Fichier 1: {file1_name}")
    c.drawString(100, height - 120, f"Fichier 2: {file2_name}")

    # Résumé du croisement
    c.drawString(100, height - 160, f"Nombre de données dans le fichier 1: {recap_df['Nombre de données fichier 1'][0]}")
    c.drawString(100, height - 180, f"Nombre de données dans le fichier 2: {recap_df['Nombre de données fichier 2'][0]}")
    c.drawString(100, height - 200, f"Résultats trouvés: {recap_df['Résultats trouvés'][0]}")
    c.drawString(100, height - 220, f"Résultats non trouvés: {recap_df['Résultats non trouvés'][0]}")

    # Détails du croisement
    c.drawString(100, height - 260, "Détails du Croisement :")
    c.setFont("Helvetica", 10)
    # Limiter à 10 lignes pour éviter que le PDF ne soit trop long
    y_position = height - 280
    for index, row in merged_df.iterrows():
        if index < 10:  # Limiter à 10 lignes
            line = f"{row.to_dict()}"  # Convertir chaque ligne en dictionnaire
            c.drawString(100, y_position, line)
            y_position -= 15  # Espacement

    c.save()
    buffer.seek(0)
    return buffer.getvalue()

def compile_excels(files, columns_to_keep):
    compiled_df = pd.DataFrame()  # DataFrame vide pour la compilation
    
    for file in files:
        df = pd.read_excel(file, dtype=str)
        df = normalize_column_names(df)
        
        # Vérification des colonnes manquantes
        missing_columns = [col for col in columns_to_keep if col not in df.columns]
        if missing_columns:
            st.warning(f"Colonnes manquantes dans {file.name}: {', '.join(missing_columns)}")
        
        # Garder seulement les colonnes qui existent
        existing_columns = [col for col in columns_to_keep if col in df.columns]
        df = df[existing_columns]
        
        # Ajouter les données compilées au DataFrame principal
        compiled_df = pd.concat([compiled_df, df], ignore_index=True)
    
    return compiled_df

def convert_df_to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)
    processed_data = output.getvalue()
    return processed_data
def convert_df_to_excel_with_formatting(df, header_color, recap_df=None):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    
    # Écrire le fichier compilé
    df.to_excel(writer, sheet_name='Compilation', index=False)
    
    # Appliquer la couleur d'en-tête
    workbook = writer.book
    header_format = workbook.add_format({'bg_color': header_color, 'bold': True})
    worksheet = writer.sheets['Compilation']
    
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    
    # Ajouter le tableau récapitulatif
    if recap_df is not None:
        recap_df.to_excel(writer, sheet_name='Récapitulatif', index=False)
        recap_worksheet = writer.sheets['Récapitulatif']
        for col_num, value in enumerate(recap_df.columns.values):
            recap_worksheet.write(0, col_num, value, header_format)
    
    writer.save()
    processed_data = output.getvalue()
    return processed_data

def cross_files(df1, df2, col_file1, col_file2):
    # Fusionner les DataFrames
    merged_df = pd.merge(df1, df2, left_on=col_file1, right_on=col_file2, how='left')
    
    # Vérifier les colonnes disponibles dans merged_df
    st.write("Colonnes après fusion :", merged_df.columns.tolist())
    
    # Assurez-vous que la colonne col_file2 existe avant d'appliquer dropna
    if col_file2 in merged_df.columns:
        merged_df = merged_df.dropna(subset=[col_file2])  # Supprimer les lignes sans correspondance
    else:
        st.warning(f"La colonne {col_file2} est introuvable dans le résultat du croisement.")
    
    return merged_df
def export_excel6(sheets_data, sheet_name):
    """
    Exporte plusieurs DataFrames dans un fichier Excel.
    
    sheets_data : liste de tuples (df, nom_onglet)
    sheet_name : Nom de l'onglet principal
    """
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')

    for df, name in sheets_data:
        if isinstance(df, pd.DataFrame):  # Vérifie si c'est bien un DataFrame
            df.to_excel(writer, sheet_name=name, index=False)

    workbook = writer.book
    for df, name in sheets_data:
        if isinstance(df, pd.DataFrame):
            worksheet = writer.sheets[name]
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.set_column(i, i, max_len + 2)

            header_format = workbook.add_format({'bold': True, 'bg_color': '#D9EAD3', 'border': 1})
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)

    writer.close()
    output.seek(0)
    return output

# Fonction pour exporter les résultats en fichier Excel avec plusieurs onglets
def export_excel(df, sheet_name, df_original=None, duplicates=None, recap_info=None, original_without_duplicates=None):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    
    # Écrire les données du DataFrame dans l'onglet spécifié
    df.to_excel(writer, sheet_name=sheet_name, index=False)

    if df_original is not None:
        df_original.to_excel(writer, sheet_name='Données Initiales', index=False)

    if duplicates is not None:
        duplicates.to_excel(writer, sheet_name='Doublons', index=False)

    if original_without_duplicates is not None:
        original_without_duplicates.to_excel(writer, sheet_name='Données Initiales sans Doublons', index=False)

    if recap_info is not None:
        recap_df = pd.DataFrame(recap_info)
        recap_df.to_excel(writer, sheet_name='Récapitulatif', index=False)

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    for i, col in enumerate(df.columns):
        max_len = max(df[col].astype(str).map(len).max(), len(col))
        worksheet.set_column(i, i, max_len + 2)

    header_format = workbook.add_format({'bold': True, 'bg_color': '#D9EAD3', 'border': 1})
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)

    writer.close()
    output.seek(0)
    return output
def export_excel2(dataframes, title):
    # Créer un buffer pour le fichier Excel
    output = io.BytesIO()
    
    # Créer un nouveau fichier Excel
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for df, sheet_name in dataframes:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Accéder à l'objet workbook et worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Définir une mise en forme pour les en-têtes
            header_format = workbook.add_format({
                'bold': True,
                'font_color': 'white',
                'bg_color': '#4F81BD',  # Couleur de fond pour les en-têtes
                'border': 1,
                'align': 'center'
            })

            # Appliquer le format aux en-têtes
            for col_num, value in enumerate(df.columns):
                worksheet.write(0, col_num, value, header_format)

            # Ajuster automatiquement la largeur des colonnes
            for i, col in enumerate(df):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2  # Ajouter un peu de marge
                worksheet.set_column(i, i, max_len)

    # Déplacer le curseur au début du buffer
    output.seek(0)
    return output.getvalue()
# Fonction pour uniformiser le format des numéros de téléphone
def uniformize_format(df, column_name):
    def clean_phone_number(phone):
        # Supprimer tous les espaces
        phone = re.sub(r'\s+', '', phone)
        # Ajouter un "0" devant s'il n'y en a pas
        if not phone.startswith('0'):
            phone = '0' + phone
        return phone

    # Appliquer la fonction de nettoyage à la colonne spécifiée
    df[column_name] = df[column_name].astype(str).apply(clean_phone_number)
    return df

# Fonction pour rechercher une valeur spécifique dans plusieurs fichiers
def search_value_in_files(files, value):
    results = []
    for file in files:
        df = pd.read_excel(file, dtype=str)
        for col in df.columns:
            if df[col].astype(str).str.contains(value, na=False).any():
                result = df[df[col].astype(str).str.contains(value, na=False)]
                results.append((file.name, col, result))
    return results

# Fonction pour rechercher les valeurs dans plusieurs fichiers
def search_multiple_values_in_files(files, values_to_search):
    results = []
    for file in files:
        try:
            # Lire chaque fichier avec tous les onglets
            excel_data = pd.read_excel(file, sheet_name=None, dtype=str)
            for sheet_name, df in excel_data.items():
                # Normaliser chaque valeur de la dataframe pour la comparer avec les valeurs recherchées
                normalized_df = df.applymap(lambda x: normalize_value(str(x)) if pd.notnull(x) else x)
                for col in normalized_df.columns:
                    # Masque de recherche pour les valeurs similaires aux valeurs recherchées
                    mask = normalized_df[col].apply(lambda x: any(x == val for val in values_to_search))
                    if mask.any():
                        # Récupérer et afficher toutes les colonnes des lignes correspondantes
                        result_data = df[mask]  # Utiliser le DataFrame d'origine pour garder les formats
                        results.append((file.name, sheet_name, col, result_data))
        except Exception as e:
            st.error(f"Erreur lors de la lecture du fichier {file.name}: {e}")
    return results
def detect_duplicates2(df, column_name):
    # Trouver les doublons dans la colonne spécifiée
    duplicates = df[df.duplicated(subset=[column_name], keep=False)]
    return duplicates
# Fonction pour lire un fichier Excel avec plusieurs onglets
def load_excel(file):
    xls = pd.ExcelFile(file)
    return xls
def export_excel4(duplicates, sheet_name, df_original=None, recap_info=None, original_without_duplicates=None):
    output = BytesIO()

    # Créer un writer avec openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Feuille des doublons
        duplicates.to_excel(writer, index=False, sheet_name=sheet_name)

        # Feuille des données initiales (si disponible)
        if df_original is not None:
            df_original.to_excel(writer, index=False, sheet_name="Données Initiales")

        # Feuille des données sans doublons (si demandée)
        if original_without_duplicates is not None:
            original_without_duplicates.to_excel(writer, index=False, sheet_name="Sans Doublons")

        # Feuille récapitulative
        if recap_info is not None:
            pd.DataFrame(recap_info).to_excel(writer, index=False, sheet_name="Récapitulatif")

        # Obtenir le workbook pour personnaliser les feuilles
        workbook = writer.book
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            # Mise en forme des en-têtes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for col_num, cell in enumerate(worksheet[1], start=1):  # Ligne 1 = En-têtes
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajuster la largeur des colonnes
            for col_num, column_cells in enumerate(worksheet.columns, start=1):
                max_length = 0
                col_letter = get_column_letter(col_num)
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)
    return output
def convert_df_to_excel_with_formatting(compiled_df, header_color, recap_df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    
    # Écriture du DataFrame compilé
    compiled_df.to_excel(writer, sheet_name='Compilé', index=False)
    
    # Récupération de l'objet workbook et de la feuille
    workbook = writer.book
    compiled_sheet = writer.sheets['Compilé']
    
    # Format d'en-tête
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': header_color,
        'font_color': 'white',
        'border': 1
    })
    
    # Appliquer le format d'en-tête
    for col_num, value in enumerate(compiled_df.columns.values):
        compiled_sheet.write(0, col_num, value, header_format)
        
    # Ajuster automatiquement la largeur des colonnes
    for i, col in enumerate(compiled_df.columns):
        max_length = max(compiled_df[col].astype(str).map(len).max(), len(col)) + 2  # Ajouter un peu de marge
        compiled_sheet.set_column(i, i, max_length)

    # Écriture du tableau récapitulatif
    recap_df.to_excel(writer, sheet_name='Récapitulatif', index=False)
    
    # Récupération de la feuille récapitulative
    recap_sheet = writer.sheets['Récapitulatif']
    
    # Appliquer le format d'en-tête pour le récapitulatif
    for col_num, value in enumerate(recap_df.columns.values):
        recap_sheet.write(0, col_num, value, header_format)
    
    # Ajuster automatiquement la largeur des colonnes pour le récapitulatif
    for i, col in enumerate(recap_df.columns):
        max_length = max(recap_df[col].astype(str).map(len).max(), len(col)) + 2  # Ajouter un peu de marge
        recap_sheet.set_column(i, i, max_length)

    writer.close()  # Utilisez close() au lieu de save()
    output.seek(0)
    return output.getvalue()
def convert_df_to_excel_with_formatting3(dfs, header_color):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Sauvegarder chaque DataFrame dans une feuille Excel différente
        for i, df in enumerate(dfs):
            df.to_excel(writer, sheet_name=f'Fichier_{i+1}', index=False)
            # Appliquer le format de l'en-tête
            workbook = writer.book
            worksheet = writer.sheets[f'Fichier_{i+1}']
            header_format = workbook.add_format({'bold': True, 'bg_color': header_color, 'align': 'center'})
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
    output.seek(0)
    return output

def detect_duplicates4(df, column_names):
    """
    Détecte les doublons en fonction des colonnes spécifiées.
    Exclut les valeurs vides et NaN avant de vérifier les doublons.
    """
    # Nettoyage des colonnes sélectionnées (supprimer les espaces, convertir en string)
    df[column_names] = df[column_names].astype(str).apply(lambda x: x.str.strip())
    
    # Exclure les valeurs NaN, vides et '0'
    df_filtered = df.dropna(subset=column_names)
    df_filtered = df_filtered[~df_filtered[column_names].isin(['', '0']).any(axis=1)]

    # Trouver les doublons
    duplicates = df_filtered[df_filtered.duplicated(subset=column_names, keep=False)]
    return duplicates
# Interface utilisateur avec Streamlit
#st.set_page_config(page_title="Application de Vérification", layout="wide")
#st.title('📊 Application de Vérification')
# Détection des doublons dans une combinaison de colonnes
def detect_combined_duplicates(df, columns):
    df_filtered = df.dropna(subset=columns)
    df_filtered = df_filtered[~df_filtered[columns].isin(['', '0']).any(axis=1)]
    combined_duplicates = df_filtered[df_filtered.duplicated(subset=columns, keep=False)]
    return combined_duplicates

# Détection des doublons dans chaque colonne
def detect_column_duplicates(df, columns):
    duplicate_dict = {}
    for col in columns:
        duplicates = df[df.duplicated(subset=[col], keep=False)]
        duplicate_dict[col] = duplicates
    return duplicate_dict

# Appliquer une mise en forme avancée aux fichiers Excel
def apply_excel_format5(writer, sheet_name, df):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    header_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        worksheet.column_dimensions[column_letter].width = max_length + 2
def clean_sheet_name(sheet_name):
    """ Nettoie le nom de l'onglet pour éviter les caractères invalides """
    invalid_chars = '[]:*?/\\'
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, "_")  # Remplace les caractères interdits
    return sheet_name[:31]  # Limite à 31 caractères (Excel impose cette limite)
# Export des données en Excel
def export_excel5(duplicate_dict, combined_duplicates, df_original, original_without_duplicates):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Feuille Données Initiales
        df_original.to_excel(writer, sheet_name="Données_Initiales", index=False)
        apply_excel_format5(writer, "Données_Initiales", df_original)

        # Feuilles des Doublons
        for col, df_dup in duplicate_dict.items():
            if not df_dup.empty:
                sheet_name = clean_sheet_name(f"Doublons_{col}")  # Nettoyer le nom de l'onglet
                df_dup.to_excel(writer, sheet_name=sheet_name, index=False)
                apply_excel_format5(writer, sheet_name, df_dup)

        # Feuille Doublons Combinés
        if not combined_duplicates.empty:
            combined_duplicates.to_excel(writer, sheet_name="Doublons_Combinés", index=False)
            apply_excel_format5(writer, "Doublons_Combinés", combined_duplicates)

        # Feuille Données Sans Doublons
        if original_without_duplicates is not None and not original_without_duplicates.empty:
            original_without_duplicates.to_excel(writer, sheet_name="Données_Sans_Doublons", index=False)
            apply_excel_format5(writer, "Données_Sans_Doublons", original_without_duplicates)

        # Feuille Récapitulatif
        recap_data = {
            "Total Lignes Initiales": [len(df_original)],
            "Total Doublons Détectés": [sum(len(df) for df in duplicate_dict.values() if not df.empty) + len(combined_duplicates)],
            "Total Après Suppression Doublons": [len(original_without_duplicates) if original_without_duplicates is not None else "Non inclus"]
        }
        recap_df = pd.DataFrame(recap_data)
        recap_df.to_excel(writer, sheet_name="Récapitulatif", index=False)
        apply_excel_format5(writer, "Récapitulatif", recap_df)

    output.seek(0)
    return output

st.markdown("""
    <style>
        /* Style général de la sidebar */
        [data-testid="stSidebar"] {
            background: linear-gradient(135deg, #1e3a8a, #9333ea);
            color: white;
        }

        /* Style du menu déroulant */
        div[data-testid="stSelectbox"] label {
            font-size: 18px;
            font-weight: bold;
            color: white;
        }

        /* Style des boutons et éléments actifs */
        .menu-button {
            display: flex;
            align-items: center;
            background: #4f46e5;
            color: white;
            padding: 10px;
            margin: 5px 0;
            border-radius: 8px;
            transition: 0.3s;
            font-weight: bold;
        }

        .menu-button:hover {
            background: #a855f7;
            transform: scale(1.05);
        }

        /* Style de l'icône */
        .menu-icon {
            margin-right: 10px;
        }

        /* Style du contenu principal */
        .main {
            background: #f3f4f6;
            padding: 20px;
            border-radius: 10px;
        }
    </style>
""", unsafe_allow_html=True)

# Définition du menu avec des icônes Unicode
menu_options = {
    "Détecteur de doublons": "🔍",
    "Croisement de fichiers": "🔗",
    "Analyse des paiements": "💰",
    "Recherche de valeur": "🔎",
    "Uniformisation des formats": "📏",
    "Compilateur de fichier": "📂",
    "Nettoyeur automatique de données": "🧹",
    "Édition Interactive": "✏️",
    "IA": "🤖",
    "Organisation de données": "📊"
}

# Sélection avec menu déroulant stylé
selected_option = st.sidebar.selectbox(
    "📌 **Choisissez une fonctionnalité**",
    options=list(menu_options.keys()),
    format_func=lambda option: f"{menu_options[option]} {option}"
)

# Affichage du contenu avec transition douce
st.markdown('<div class="main">', unsafe_allow_html=True)

st.header(f"{menu_options[selected_option]} {selected_option}")  


# --- Interface Streamlit ---
# --- Interface Streamlit ---
if selected_option == "Détecteur de doublons":

    uploaded_file = st.file_uploader("Choisissez un fichier Excel", type="xlsx")
    if uploaded_file is not None:
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names
        selected_sheet = st.selectbox("Choisissez un onglet", sheet_names)
        df = pd.read_excel(xls, sheet_name=selected_sheet, dtype=str)
        st.write("Aperçu des données :", df.head())
        column_names = st.multiselect("Choisissez les colonnes pour détecter les doublons", df.columns)
        include_original_without_duplicates = st.checkbox("Inclure les données initiales sans doublons dans l'export")
        if st.button("Détecter les doublons"):
            if not column_names:
                st.warning("Veuillez sélectionner au moins une colonne.")
            else:
                with st.spinner("Détection des doublons en cours..."):
                    time.sleep(2)
                    duplicate_dict = detect_column_duplicates(df.copy(), column_names)
                    combined_duplicates = detect_combined_duplicates(df.copy(), column_names)
                    all_duplicates_indices = set()
                    for df_dup in duplicate_dict.values():
                        all_duplicates_indices.update(df_dup.index)
                    if not combined_duplicates.empty:
                        all_duplicates_indices.update(combined_duplicates.index)
                    original_without_duplicates = None
                    if include_original_without_duplicates:
                        original_without_duplicates = df.drop_duplicates(subset=column_names, keep='first')
                    total_doublons = sum(len(df) for df in duplicate_dict.values() if not df.empty) + len(combined_duplicates)
                    total_sans_doublons = len(original_without_duplicates) if original_without_duplicates is not None else "Non inclus"
                    if total_doublons > 0:
                        st.error("⚠️ Doublons detectés !")
                        st.write(f"### Récapitulatif")
                        st.write(f"- Nombre total de lignes initiales : {len(df)}")
                        st.write(f"- Nombre total de doublons détectés : {total_doublons}")
                        st.write(f"- Nombre total après suppression des doublons : {total_sans_doublons}")
                        for col, duplicates in duplicate_dict.items():
                            if not duplicates.empty:
                                st.write(f"### Doublons dans la colonne **{col}** :", duplicates)
                        if not combined_duplicates.empty:
                            st.write("### Doublons combinés sur les colonnes sélectionnées :", combined_duplicates)
                        excel_data = export_excel5(
                            duplicate_dict,
                            combined_duplicates,
                            df_original=df,
                            original_without_duplicates=original_without_duplicates
                        )
                        st.download_button(
                            label="Télécharger les doublons en Excel",
                            data=excel_data,
                            file_name="doublons.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.success("✅ Aucun doublon detecté.")


elif selected_option == "Croisement de fichiers":
 

    # Charger les deux fichiers Excel pour le croisement
    file1 = st.file_uploader("Choisissez le premier fichier Excel", type="xlsx", key="file1")
    file2 = st.file_uploader("Choisissez le deuxième fichier Excel", type="xlsx", key="file2")

    if file1 is not None and file2 is not None:
        xls1 = load_excel(file1)
        xls2 = load_excel(file2)

        # Sélectionner les onglets pour chaque fichier
        selected_sheet1 = st.selectbox("Choisissez un onglet du fichier 1", xls1.sheet_names, key="sheet1")
        selected_sheet2 = st.selectbox("Choisissez un onglet du fichier 2", xls2.sheet_names, key="sheet2")

        df1 = pd.read_excel(xls1, sheet_name=selected_sheet1, dtype=str)
        df2 = pd.read_excel(xls2, sheet_name=selected_sheet2, dtype=str)

        st.write("Aperçu du fichier 1 :", df1.head())
        st.write("Aperçu du fichier 2 :", df2.head())

        # Choisir la colonne pour le croisement
        col_file1 = st.selectbox("Choisissez la colonne du fichier 1 pour le croisement", df1.columns, key="col_file1")
        col_file2 = st.selectbox("Choisissez la colonne du fichier 2 pour le croisement", df2.columns, key="col_file2")

        # Type de jointure
        join_type = st.selectbox("Type de jointure", ["left", "right", "outer", "inner"], key="join_type")

        # Filtrage des données
        filter_value = st.text_input("Filtrer les données du fichier 1 par cette valeur (laisser vide pour tout afficher)", "")
        if filter_value:
            df1 = df1[df1[col_file1].str.contains(filter_value, na=False)]

        if st.button("Croiser les fichiers", key="cross_files"):
            with st.spinner("Croisement des fichiers en cours..."):
                time.sleep(2)  # Simuler un chargement
                
                # Réaliser le croisement
                merged_df = df1.merge(df2[[col_file2]], left_on=col_file1, right_on=col_file2, how=join_type)
                merged_df['Résultat de croisement'] = merged_df[col_file2].notnull().replace({True: 'Trouvé', False: 'Non Trouvé'})

                st.success("Croisement terminé !")

            # Afficher le résultat du croisement
            st.write("Résultat du croisement :", merged_df.head())

            # Tableau récapitulatif
            recap_info = {
                'Nombre de données fichier 1': [len(df1)],
                'Nombre de données fichier 2': [len(df2)],
                'Résultats trouvés': [merged_df['Résultat de croisement'].value_counts().get('Trouvé', 0)],
                'Résultats non trouvés': [merged_df['Résultat de croisement'].value_counts().get('Non Trouvé', 0)]
            }

            recap_df = pd.DataFrame(recap_info)
            st.write("Tableau récapitulatif :")
            st.write(recap_df)

            # Statistiques descriptives
            st.write("### Statistiques descriptives :")
            st.write(merged_df.describe())

            # Créer des graphiques pour visualiser les résultats
            st.bar_chart(recap_df.set_index(recap_df.columns[0]))

            # Explication des résultats
            st.write("### Explications sur le croisement :")
            st.write(f"- Le fichier 1 contient **{len(df1)}** enregistrements.")
            st.write(f"- Le fichier 2 contient **{len(df2)}** enregistrements.")
            st.write(f"- Après le croisement, nous avons trouvé **{recap_df.iloc[0, 2]}** enregistrements dans le fichier 1 qui correspondent aux enregistrements du fichier 2.")
            st.write(f"- **{recap_df.iloc[0, 3]}** enregistrements n'ont pas été trouvés dans le fichier 2.")

            # Initialiser les colonnes sélectionnées dans session_state si non existantes
            if 'selected_columns' not in st.session_state:
                st.session_state.selected_columns = merged_df.columns.tolist()

            # Sélection des colonnes à exporter
            selected_columns = st.multiselect("Choisissez les colonnes à exporter", merged_df.columns.tolist(), default=st.session_state.selected_columns)

            # Sauvegarder la sélection
            st.session_state.selected_columns = selected_columns

            if selected_columns:
                # Préparation des données pour l'exportation
                sheets_data = [
                    (merged_df[selected_columns], 'Croisement'),
                    (recap_df, 'Récapitulatif')
                ]
                
                # Génération du fichier Excel
                excel_data = export_excel6(sheets_data, 'Croisement')
            
                # Téléchargement du fichier Excel
                st.download_button(
                    label="Télécharger le fichier croisé en Excel",
                    data=excel_data,
                    file_name="fichier_croise.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

elif selected_option == "Analyse des paiements":

    
    payment_file = st.file_uploader("Choisissez le fichier des paiements", type="xlsx", key="payment_file")
    transaction_file = st.file_uploader("Choisissez le fichier des transactions", type="xlsx", key="transaction_file")

    if payment_file is not None and transaction_file is not None:
        payments_df = pd.read_excel(payment_file, dtype=str)
        transactions_df = pd.read_excel(transaction_file, dtype=str)
        st.write("Aperçu du fichier des paiements :", payments_df.head())
        st.write("Aperçu du fichier des transactions :", transactions_df.head())

        col_payment = st.selectbox("Choisissez la colonne pour les paiements", payments_df.columns)
        col_transaction = st.selectbox("Choisissez la colonne pour les transactions", transactions_df.columns)

        # Ajout d'un selectbox pour choisir une colonne supplémentaire à afficher
        additional_column = st.selectbox("Choisissez une colonne supplémentaire à afficher", transactions_df.columns, index=0)

        if st.button("Analyser les paiements"):
            merged_payments = pd.merge(payments_df, transactions_df, left_on=col_payment, right_on=col_transaction, how='left')

            # Compter les paiements
            count_payments = merged_payments[col_transaction].value_counts().reset_index()
            count_payments.columns = [col_transaction, 'Nombre de fois payé']

            # Ajouter la colonne supplémentaire au DataFrame
            count_payments = count_payments.merge(merged_payments[[col_transaction, additional_column]], on=col_transaction, how='left')

            st.write("Analyse des paiements :", count_payments)

            excel_data = export_excel(count_payments, 'Analyse des paiements')
            st.download_button(
                label="Télécharger l'analyse des paiements en Excel",
                data=excel_data,
                file_name="analyse_paiements.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.xlsx"
            )

elif selected_option == "Recherche de valeur":
   
    search_files = st.file_uploader("Choisissez un ou plusieurs fichiers Excel pour la recherche", type="xlsx", accept_multiple_files=True, label_visibility="collapsed")

    if search_files:
        search_values = st.text_area("Entrez les valeurs à rechercher (peuvent être séparées par des espaces, des retours à la ligne, des virgules, etc.)")

        if st.button("Rechercher", key="search_values"):
            if search_values:
                # Extraire et normaliser les valeurs de recherche
                values_to_search = extract_values(search_values)

                with st.spinner("Recherche en cours..."):
                    time.sleep(2)  # Simuler un chargement
                    search_results = search_multiple_values_in_files(search_files, values_to_search)

                if search_results:
                    for file_name, sheet_name, col, result_data in search_results:
                        st.write(f"Résultats trouvés dans {file_name} - Onglet {sheet_name} - Colonne {col}:")
                        st.dataframe(result_data)  # Affiche toutes les colonnes pour les lignes correspondantes
                else:
                    st.warning("Aucun résultat trouvé pour les valeurs recherchées.")
            else:
                st.warning("Veuillez entrer au moins une valeur à rechercher.")

elif selected_option == "Uniformisation des formats":


    # Charger un fichier Excel
    excel_file = st.file_uploader("Choisissez le fichier Excel à uniformiser", type="xlsx")

    if excel_file is not None:
        # Lire le fichier et afficher un aperçu
        df = pd.read_excel(excel_file, dtype=str)
        st.write("Aperçu des données :", df.head())

        # Initialiser un dictionnaire pour les noms de colonnes modifiés
        modified_columns = {}

        # Uniformisation des en-têtes de colonnes
        if st.checkbox("Uniformiser les en-têtes de colonnes"):
            unique_columns = df.columns.tolist()
            st.write("Noms de colonnes détectés :")
            
            # Dictionnaire pour stocker les nouveaux noms de colonnes
            modified_columns = {}
            
            for col in unique_columns:
                new_col_name = st.text_input(f"Modifier le nom de la colonne '{col}' :", value=col)
                modified_columns[col] = new_col_name

            # Appliquer l'uniformisation des en-têtes
            df = uniformize_headers(df, modified_columns)

        # Uniformisation des formats
        column_to_uniformize_format = None
        if st.checkbox("Uniformiser les numero mobiles"):
            column_to_uniformize_format = st.selectbox("Choisissez la colonne à uniformiser les formats", df.columns)

        # Uniformisation des textes
        text_column = None
        if st.checkbox("Uniformiser les textes (enlever les sur-espaces)"):
            text_column = st.selectbox("Choisissez la colonne pour uniformiser les textes", df.columns)

        # Vérification de l'existence des colonnes avant d'appliquer les transformations
        if st.button("Appliquer les uniformisations"):
            with st.spinner("Uniformisation en cours..."):
                time.sleep(2)  # Simuler un chargement

                # Appliquer l'uniformisation des formats
                if column_to_uniformize_format and column_to_uniformize_format in df.columns:
                    df = uniformize_format(df, column_to_uniformize_format)

                # Appliquer l'uniformisation des textes
                if text_column and text_column in df.columns:
                    df = uniformize_text(df, text_column)

                st.success("Uniformisation terminée !")

                # Afficher les données uniformisées
                st.write("Données uniformisées :", df.head())
                excel_data = export_excel(df, 'Données Uniformisées')
                st.download_button(
                    label="Télécharger les données uniformisées en Excel",
                    data=excel_data,
                    file_name="donnees_uniformisees.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
elif selected_option == "Compilateur de fichier" :


    # Sélectionner entre l'assemblage de plusieurs fichiers ou de plusieurs onglets dans un même fichier
    compilation_type = st.radio("Choisissez le type de compilation :", 
                                options=["Assembler plusieurs fichiers Excel", "Assembler plusieurs onglets d'un même fichier"])

    if compilation_type == "Assembler plusieurs fichiers Excel":
        # Charger plusieurs fichiers Excel
        excel_files = st.file_uploader("Choisissez les fichiers Excel à compiler", type="xlsx", accept_multiple_files=True)
    
        if excel_files:
            all_columns = []
            file_details = []
            for file in excel_files:
                try:
                    df = pd.read_excel(file, dtype=str)
                    df = normalize_column_names(df)  # Normaliser les noms de colonnes
                    all_columns.extend(df.columns.tolist())
                    file_details.append({'file_name': file.name, 'row_count': len(df)})
                except Exception as e:
                    st.error(f"Erreur lors de la lecture du fichier {file.name}: {e}")
            
            all_columns = list(set(all_columns))  # Supprimer les doublons de colonnes
            selected_columns = st.multiselect("Choisissez les colonnes à compiler", all_columns)
            header_color = st.color_picker("Choisissez une couleur pour l'entête", "#00A6D6")

            if selected_columns:
                if st.button("Compiler les fichiers"):
                    with st.spinner("Compilation en cours..."):
                        compiled_df = compile_excels(excel_files, selected_columns)
                        st.success("Compilation terminée !")

                        recap_data = [{'Nom du fichier': f['file_name'], 'Nombre de lignes': f['row_count']} for f in file_details]
                        total_rows = sum(f['row_count'] for f in file_details)
                        recap_data.append({'Nom du fichier': 'Sous-total', 'Nombre de lignes': total_rows})
                        recap_df = pd.DataFrame(recap_data)

                        st.write("Aperçu du fichier compilé :", compiled_df.head())
                        st.write("Tableau récapitulatif des fichiers compilés :")
                        st.write(recap_df)

                        compiled_excel = convert_df_to_excel_with_formatting(compiled_df, header_color, recap_df)
                        
                        st.download_button(
                            label="Télécharger le fichier compilé en Excel",
                            data=compiled_excel,
                            file_name="fichier_compilé_formaté.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.warning("Veuillez sélectionner au moins une colonne.")

    elif compilation_type == "Assembler plusieurs onglets d'un même fichier":
        excel_file = st.file_uploader("Choisissez le fichier Excel contenant plusieurs onglets", type="xlsx", accept_multiple_files=False)

        if excel_file:
            # Charger les onglets du fichier Excel
            dfs = pd.read_excel(excel_file, sheet_name=None, dtype=str)
            all_sheets = list(dfs.keys())
            selected_sheets = st.multiselect("Choisissez les onglets à compiler", all_sheets)

            if selected_sheets:
                # Extraire toutes les colonnes disponibles à partir des onglets sélectionnés
                all_columns = []
                for sheet in selected_sheets:
                    df = dfs[sheet]
                    df = normalize_column_names(df)  # Normaliser les noms de colonnes
                    all_columns.extend(df.columns.tolist())
                
                all_columns = list(set(all_columns))  # Supprimer les doublons de colonnes
                selected_columns = st.multiselect("Choisissez les colonnes à compiler", all_columns)
                header_color = st.color_picker("Choisissez une couleur pour l'entête", "#00A6D6")

                if selected_columns:
                    if st.button("Compiler les onglets"):
                        with st.spinner("Compilation en cours..."):
                            compiled_dfs = [
                                            dfs[sheet][dfs[sheet].columns.intersection(selected_columns)]
                                            for sheet in selected_sheets
                                        ]
                            compiled_df = pd.concat(compiled_dfs, ignore_index=True)
                            st.success("Compilation des onglets terminée !")

                            recap_data = [{'Nom de l\'onglet': sheet, 'Nombre de lignes': len(dfs[sheet])} for sheet in selected_sheets]
                            total_rows = sum(len(dfs[sheet]) for sheet in selected_sheets)
                            recap_data.append({'Nom de l\'onglet': 'Sous-total', 'Nombre de lignes': total_rows})
                            recap_df = pd.DataFrame(recap_data)

                            st.write("Aperçu du fichier compilé :", compiled_df.head())
                            st.write("Tableau récapitulatif des onglets compilés :")
                            st.write(recap_df)

                            compiled_excel = convert_df_to_excel_with_formatting(compiled_df, header_color, recap_df)
                            
                            st.download_button(
                                label="Télécharger le fichier compilé en Excel",
                                data=compiled_excel,
                                file_name="onglets_compilés_formaté.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                else:
                    st.warning("Veuillez sélectionner au moins une colonne.")

elif selected_option == "Nettoyeur automatique de données":

    
    # Option de nettoyage automatique
    nettoyage_auto = st.checkbox("Activer le nettoyage automatique des données", key="nettoyage_auto")

    # Charger plusieurs fichiers Excel avec des clés uniques
    excel_files = st.file_uploader("Choisissez les fichiers Excel à compiler", type="xlsx", accept_multiple_files=True, key="file_uploader_compilation")

    if excel_files:
        all_columns = []
        file_details = []
        compiled_dataframes = []

        for file in excel_files:
            try:
                df = pd.read_excel(file, dtype=str)
                
                # Appliquer le nettoyage automatique si activé
                if nettoyage_auto:
                    df = nettoyer_donnees(df)
                
                all_columns.extend(df.columns.tolist())
                file_details.append({'file_name': file.name, 'row_count': len(df)})
                compiled_dataframes.append(df)
                
            except Exception as e:
                st.error(f"Erreur lors de la lecture du fichier {file.name}: {e}")

        # Supprimer les doublons de colonnes
        all_columns = list(set(all_columns))

        # Permettre à l'utilisateur de choisir les colonnes à compiler
        selected_columns = st.multiselect("Choisissez les colonnes à compiler", all_columns, key="columns_selection_compilation")
        
        # Choisir la couleur de l'en-tête
        header_color = st.color_picker("Choisissez une couleur pour l'entête", "#00A6D6", key="header_color_compilation")

        if selected_columns and compiled_dataframes:
            if st.button("Compiler les fichiers", key="compile_files"):
                with st.spinner("Compilation en cours..."):
                    compiled_df = pd.concat([df[selected_columns] for df in compiled_dataframes], ignore_index=True)
                    st.success("Compilation terminée !")
                    
                    # Créer un tableau récapitulatif
                    recap_data = [{'Nom du fichier': f['file_name'], 'Nombre de lignes': f['row_count']} for f in file_details]
                    total_rows = sum(f['row_count'] for f in file_details)
                    recap_data.append({'Nom du fichier': 'Sous-total', 'Nombre de lignes': total_rows})
                    recap_df = pd.DataFrame(recap_data)

                    # Afficher un aperçu du fichier compilé et du récapitulatif
                    st.write("Aperçu du fichier compilé :", compiled_df.head())
                    st.write("Tableau récapitulatif des fichiers compilés :")
                    st.write(recap_df)

                    # Exporter le fichier compilé avec mise en forme
                    compiled_excel = convert_df_to_excel_with_formatting(compiled_df, header_color, recap_df)
                    
                    st.download_button(
                        label="Télécharger le fichier compilé en Excel",
                        data=compiled_excel,
                        file_name="fichier_compilé_nettoyé.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_compiled_file"
                    )
        else:
            st.warning("Veuillez sélectionner au moins une colonne.")
elif selected_option == "Édition Interactive":


    # Charger le fichier Excel pour modification
# Charger le fichier Excel pour modification
    uploaded_file = st.file_uploader("Choisissez un fichier Excel", type="xlsx", key="uploader1")

    if uploaded_file:
        # Lire le fichier Excel
        df = pd.read_excel(uploaded_file)

        # Affichage des données dans une grille interactive
        st.write("Modifiez les données dans la table ci-dessous :")
        
        # Configuration de la grille interactive
        gb = GridOptionsBuilder.from_dataframe(df)
        gb.configure_pagination(paginationAutoPageSize=True)  # Pagination automatique
        gb.configure_side_bar()  # Barre latérale pour les filtres
        gb.configure_default_column(editable=True)  # Rendre toutes les colonnes éditables
        
        grid_options = gb.build()

        # Affichage de la grille interactive avec AgGrid
        grid_response = AgGrid(
            df,
            gridOptions=grid_options,
            update_mode=GridUpdateMode.MODEL_CHANGED,  # Met à jour lorsque le modèle change
            editable=True
        )

        # Obtenir les données modifiées
        modified_data = grid_response['data']
        df_modified = pd.DataFrame(modified_data)

        # Afficher l'aperçu des données modifiées
        st.write("Aperçu des données modifiées :", df_modified)

        # Permettre à l'utilisateur de télécharger le fichier modifié
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df_modified.to_excel(writer, index=False)
            writer.close()

        st.download_button(
            label="Télécharger les données modifiées",
            data=buffer,
            file_name="données_modifiées.xlsx",
            mime="application/vnd.ms-excel"
        )
elif selected_option == "IA":


    # Charger le fichier de données principal
    uploaded_file = st.file_uploader("Choisissez un fichier Excel (Données)", type="xlsx", key="file_uploader_donnees")

    # Charger le fichier de transactions
    uploaded_transactions = st.file_uploader("Choisissez un fichier Excel (Transactions)", type="xlsx", key="file_uploader_transactions")

    # Vérifier que les deux fichiers sont chargés
    if uploaded_file and uploaded_transactions:
        # Lire les fichiers Excel
        df = pd.read_excel(uploaded_file)
        df_transactions = pd.read_excel(uploaded_transactions)

        # Afficher les données pour l'utilisateur
        st.write("Données chargées:")
        st.write(df.head())

        st.write("Transactions chargées:")
        st.write(df_transactions.head())

        # Section du Chatbot
        st.sidebar.subheader("Chat avec IA")

        # Initialiser l'historique des conversations
        if "conversation" not in st.session_state:
            st.session_state.conversation = []

        # Ajouter un bouton pour afficher ou masquer le chat
        show_chat = st.sidebar.checkbox("Afficher le chat", value=False)

        if show_chat:
            # Interface de chat
            user_input = st.sidebar.text_input("Vous : ", key="user_input")

            # Lorsqu'un message est envoyé
            if st.sidebar.button("Envoyer"):
                # Ajouter la question de l'utilisateur à la conversation
                st.session_state.conversation.append(f"Vous : {user_input}")

                # Obtenir la réponse du chatbot
                response = chatbot_response(user_input, df, df_transactions)

                # Ajouter la réponse du chatbot à la conversation
                st.session_state.conversation.append(f"IA : {response}")

            # Afficher l'historique des conversations
            st.sidebar.subheader("Historique des conversations")
            for message in st.session_state.conversation:
                st.sidebar.write(message)

        # Affichage des données dans une grille interactive (modifiable)
        st.write("Modifiez les données dans la table ci-dessous si nécessaire :")
        
        grid_response = AgGrid(
            df,
            update_mode=GridUpdateMode.MODEL_CHANGED,
            editable=True
        )

        modified_data = grid_response['data']
        df_modified = pd.DataFrame(modified_data)

        # Téléchargement des données modifiées
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df_modified.to_excel(writer, index=False)
            writer.close()

        st.download_button(
            label="Télécharger les données modifiées",
            data=buffer,
            file_name="données_modifiées.xlsx",
            mime="application/vnd.ms-excel",
            key="download_button_donnees_modifiees"  # Ajoutez une clé unique ici
        )
elif selected_option == "Organisation de données":
  

    # Charger le fichier de données principal
    uploaded_file = st.file_uploader("Choisissez un fichier Excel à organiser", type="xlsx", key="file_uploader_organize")

    if uploaded_file:
        # Lire le fichier Excel
        df = pd.read_excel(uploaded_file)

        # Afficher les données originales
        st.subheader("Données Originales")
        st.write(df)

        # Organiser les données automatiquement
        organized_df = organize_data(df)

        # Afficher les données organisées
        st.subheader("Données Organisées")
        st.write(organized_df)

        # Ajouter une table interactive modifiable pour les données organisées
        st.subheader("Modifiez les données organisées si nécessaire :")
        gb = GridOptionsBuilder.from_dataframe(organized_df)
        gb.configure_default_column(editable=True)  # Rendre toutes les colonnes modifiables
        grid_options = gb.build()
        grid_response = AgGrid(
            organized_df,
            gridOptions=grid_options,
            update_mode=GridUpdateMode.MODEL_CHANGED,
            editable=True,
            height=400,
        )

        modified_data = grid_response['data']
        df_modified = pd.DataFrame(modified_data)

        # Télécharger les données organisées modifiées
        st.subheader("Télécharger les Données Modifiées")
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df_modified.to_excel(writer, index=False)
            writer.close()

        st.download_button(
            label="Télécharger le fichier Excel organisé",
            data=buffer,
            file_name="donnees_organisees.xlsx",
            mime="application/vnd.ms-excel",
            key="download_button_organized"
        )

